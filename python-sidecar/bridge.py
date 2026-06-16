"""GBT Register Manager - Python Sidecar bridge.

协议（与 Rust 侧保持一致）：
- stdin  : UTF-8 JSON 单帧（Rust 写入后立即 close，收到 EOF 开始处理）
- stdout : UTF-8 文本，业务结果包裹在 `<<<GBT-BEGIN>>>` / `<<<GBT-END>>>` 之间；
           SDK/第三方库的任意垃圾输出都不会污染结果。
- stderr : 调试日志（`[GBT-PY] ...`），Rust 侧不用于解析。

退出码：
- 0 : 已输出一帧结果（ok 字段为 true 或 false）
- 1 : 未输出帧（sidecar 自身崩溃，Rust 兜底构造错误）
"""

from __future__ import annotations

import json
import logging
import math
import os
import socket
import sys
import traceback
from contextlib import contextmanager
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Callable, Dict, Iterator, List, Optional, Set, Tuple

# ---------- 环境准备 --------------------------------------------------------

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", newline="")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace", newline="")
# 协议帧走原始 fd 直写，避免子进程管道全缓冲导致进度事件批量到达前端。
_PROTOCOL_FD = sys.stdout.fileno()

try:
    import openpyxl  # noqa: F401  (显式校验，缺失时给出清晰错误)
    from openpyxl import Workbook
except Exception as _openpyxl_err:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    _OPENPYXL_ERR = _openpyxl_err
else:
    _OPENPYXL_ERR = None

# ---------- 日志 ------------------------------------------------------------


def _resolve_log_dir() -> Path:
    """优先使用 Rust 通过环境变量传入的 app_log_dir；否则回退到用户目录。"""
    env_dir = os.environ.get("GBT_LOG_DIR")
    if env_dir:
        try:
            p = Path(env_dir)
            p.mkdir(parents=True, exist_ok=True)
            return p
        except Exception:
            pass
    fallback = Path.home() / ".gbt-register-manager" / "logs"
    try:
        fallback.mkdir(parents=True, exist_ok=True)
    except Exception:
        fallback = Path(os.getcwd())
    return fallback


_LOG_DIR = _resolve_log_dir()
_LOGGER = logging.getLogger("gbt-py")
_LOGGER.setLevel(logging.INFO)
_LOGGER.propagate = False
if not _LOGGER.handlers:
    _stderr_handler = logging.StreamHandler(sys.stderr)
    _stderr_handler.setFormatter(logging.Formatter("[GBT-PY] %(asctime)s %(levelname)s %(message)s"))
    _LOGGER.addHandler(_stderr_handler)
    try:
        _file_handler = RotatingFileHandler(
            _LOG_DIR / "gbt-py.log", maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"
        )
        _file_handler.setFormatter(
            logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
        )
        _LOGGER.addHandler(_file_handler)
    except Exception:  # pragma: no cover
        pass


def log_py(msg: str) -> None:
    _LOGGER.info(msg)


# ---------- Agilebot SDK 延迟加载 -------------------------------------------

try:
    from Agilebot import (
        Arm,
        PoseRegister,
        PoseType,
        Posture,
        ProgramPose,
        StatusCodeEnum,
    )
except Exception as _sdk_err:  # pragma: no cover
    Arm = None  # type: ignore[assignment]
    PoseRegister = None  # type: ignore[assignment]
    PoseType = None  # type: ignore[assignment]
    Posture = None  # type: ignore[assignment]
    ProgramPose = None  # type: ignore[assignment]
    StatusCodeEnum = None  # type: ignore[assignment]
    _SDK_IMPORT_ERR: Exception | None = _sdk_err
else:
    _SDK_IMPORT_ERR = None


def _ensure_sdk_ready() -> None:
    if Arm is None or StatusCodeEnum is None:
        if _SDK_IMPORT_ERR is not None:
            log_py(f"SDK import failed: {_SDK_IMPORT_ERR!r}")
        gbt_raise("GBT_SDK_NOT_FOUND")


# ---------- 协议辅助 --------------------------------------------------------

FRAME_BEGIN = "<<<GBT-BEGIN>>>"
FRAME_END = "<<<GBT-END>>>"


def gbt_raise(code: str) -> None:
    """抛出仅含 `GBT_*` 错误码的异常，供 main() 写入 sidecar 帧。"""
    raise RuntimeError(code)


def sidecar_error_code(exc: BaseException) -> str:
    msg = str(exc).strip()
    if msg.startswith("GBT_") and " " not in msg:
        return msg
    return "GBT_INTERNAL_ERROR"


def emit_frame(data: Dict[str, Any]) -> None:
    """写出被分隔符包裹的单帧 JSON（UTF-8，不转义非 ASCII）。"""
    text = json.dumps(data, ensure_ascii=False, allow_nan=False, separators=(",", ":"))
    chunk = f"{FRAME_BEGIN}\n{text}\n{FRAME_END}\n".encode("utf-8")
    os.write(_PROTOCOL_FD, chunk)


ProgressCallback = Callable[[int, Optional[int], int], None]


def emit_progress(action: str, current: int, total: Optional[int], matched: int) -> None:
    emit_frame(
        {
            "kind": "progress",
            "action": action,
            "current": current,
            "total": total,
            "matched": matched,
        }
    )


def read_payload() -> Dict[str, Any]:
    """从 stdin 读取 UTF-8 JSON 载荷（Rust 端写完后 close stdin）。"""
    raw = sys.stdin.read()
    if not raw:
        raise RuntimeError("GBT_INTERNAL_ERROR")
    return json.loads(raw)


@contextmanager
def redirect_sdk_stdout_to_stderr() -> Iterator[None]:
    """保护 stdout 不被 SDK/第三方库污染——它们 print 的一切重定向到 stderr。"""
    original = sys.stdout
    sys.stdout = sys.stderr
    try:
        yield
    finally:
        sys.stdout = original


# ---------- 通用工具 --------------------------------------------------------


def round3(v: Any) -> float:
    x = float(v)
    if math.isnan(x) or math.isinf(x):
        return 0.0
    return round(x, 3)


def unwrap_status(ret: Any) -> Any:
    if isinstance(ret, tuple) and len(ret) >= 2:
        return ret[-1]
    return ret


def status_text(ret: Any) -> str:
    st = unwrap_status(ret)
    return getattr(st, "errmsg", str(st))


def status_indicates_exists(ret: Any) -> bool:
    st = unwrap_status(ret)
    raw = " ".join(
        str(x)
        for x in (
            getattr(st, "name", ""),
            getattr(st, "value", ""),
            getattr(st, "errmsg", ""),
            st,
        )
        if x is not None
    ).lower()
    return any(
        token in raw
        for token in (
            "exist",
            "exists",
            "already",
            "duplicate",
            "conflict",
            "已存在",
            "已经存在",
            "存在",
            "重複",
            "重复",
        )
    )


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _coord_from_pose(pose: Any) -> str:
    try:
        lr = pose.poseData.cartData.baseCart.posture.arm_left_right
    except Exception:
        return "R"
    if isinstance(lr, str):
        s = lr.strip().upper()
        if s in ("L", "LEFT"):
            return "L"
        if s in ("R", "RIGHT"):
            return "R"
    n = _safe_int(lr, 1)
    return "L" if n < 0 else "R"


def _coord_to_left_right(coord: str) -> int:
    """Excel 里 L/R 映射到 SDK `arm_left_right`：L → -1，R/其他 → 1。"""
    c = str(coord or "").strip().upper()
    return -1 if c == "L" else 1


def _new_posture(coord: str) -> Any:
    """按 SDK PR/P 示例构造 Posture：显式 new 一个实例并写入需要的字段。

    避免对未初始化 posture 做属性链赋值造成的 "静默丢失"。
    对 SDK 未显式给默认值的 `arm_back_front` 一并补成 1（示例取值），
    保证机器人侧拿到可用的完整姿态。
    """
    if Posture is None:
        return None
    posture = Posture()
    try:
        posture.arm_left_right = _coord_to_left_right(coord)
    except Exception:
        pass
    try:
        # 与 PR.py 示例保持一致；若 SDK 对该字段有其它取值也不会抛异常。
        posture.arm_back_front = 1
    except Exception:
        pass
    return posture


def _apply_coord_to_pose(pose: Any, coord: str) -> None:
    """把 Excel 里的 coord (L/R) 写到 P 点的 `baseCart.posture.arm_left_right`。

    如果 `posture` 字段本身为空/不可写，则 new 一个 Posture 挂上去，
    避免静默丢失（见 SDK program_pose.py 和 PR.py 示例）。
    """
    value = _coord_to_left_right(coord)
    cart_data = pose.poseData.cartData.baseCart
    try:
        cart_data.posture.arm_left_right = value
        return
    except Exception:
        pass
    try:
        cart_data.posture = _new_posture(coord)
    except Exception:
        log_py("_apply_coord_to_pose 无法设置 posture，保留 SDK 默认值")


def _coord_from_pose_register(pose_register: Any) -> str:
    try:
        lr = pose_register.poseRegisterData.cartData.posture.arm_left_right
    except Exception:
        return "R"
    if isinstance(lr, str):
        s = lr.strip().upper()
        if s in ("L", "LEFT"):
            return "L"
        if s in ("R", "RIGHT"):
            return "R"
    n = _safe_int(lr, 1)
    return "L" if n < 0 else "R"


def _apply_coord_to_pose_register(pose_register: Any, coord: str) -> None:
    """把 Excel 里的 coord (L/R) 写到 PR 寄存器 `cartData.posture.arm_left_right`。

    PR.py 示例明确要求：新建 PoseRegister 后需 new 一个 Posture 再挂到
    `poseRegisterData.cartData.posture`，否则属性链写入会静默丢失。
    """
    value = _coord_to_left_right(coord)
    cart_data = pose_register.poseRegisterData.cartData
    try:
        cart_data.posture.arm_left_right = value
        return
    except Exception:
        pass
    try:
        cart_data.posture = _new_posture(coord)
    except Exception:
        log_py("_apply_coord_to_pose_register 无法设置 posture，保留 SDK 默认值")


def make_headers(register_type: str) -> List[str]:
    if register_type == "R":
        return ["type", "ID", "value"]
    if register_type == "P":
        return ["Type", "ID", "X", "Y", "Z", "A", "B", "C", "TF", "UF", "Coord"]
    if register_type == "PR":
        return ["TYPE", "ID", "X", "Y", "Z", "A", "B", "C", "coord"]
    raise ValueError("GBT_UNSUPPORTED_REGISTER_TYPE")


# P 寄存器（程序位姿）服务的 HTTP 端口，与 R/PR 所用端口不同。
P_SERVICE_PORT = 5606

# 「读取全部」：从该 ID 起顺序尝试，连续失败达到上限则停止。
READ_ALL_FIRST_ID = 1
READ_ALL_CONSECUTIVE_FAIL_LIMIT = 10
READ_ALL_MAX_ID = 100000


def _normalize_id(value: Any) -> Optional[int]:
    try:
        idx = int(value)
    except Exception:
        return None
    if idx < READ_ALL_FIRST_ID or idx > READ_ALL_MAX_ID:
        return None
    return idx


def build_indexes(selector: Dict[str, Any]) -> List[int]:
    mode = selector.get("mode", "range")
    if mode == "all":
        raise RuntimeError("GBT_INTERNAL_ERROR")
    start_id = int(selector.get("startId", 0))
    end_id = int(selector.get("endId", start_id))
    if end_id < start_id:
        start_id, end_id = end_id, start_id
    return list(range(start_id, end_id + 1))


# ---------- SDK 交互 --------------------------------------------------------


def _extract_connect_params(payload: Dict[str, Any]) -> Dict[str, Any]:
    """从 payload 中抽取连接参数，兼容新旧调用。"""
    controller_ip = (
        payload.get("controllerIp")
        or payload.get("ip")
        or ""
    )
    teach_panel_ip = payload.get("teachPanelIp") or None
    if isinstance(teach_panel_ip, str) and not teach_panel_ip.strip():
        teach_panel_ip = None
    local_proxy = bool(payload.get("localProxy", False))
    return {
        "controller_ip": controller_ip,
        "teach_panel_ip": teach_panel_ip,
        "local_proxy": local_proxy,
    }


def connect_arm(controller_ip: str, teach_panel_ip: Any = None, local_proxy: bool = False):
    """连接机器人（SDK 4.1.1）。

    - controller_ip: 控制柜 IP（必填）
    - teach_panel_ip: 示教器 IP（可选，仅工业机器人常用；协作/四轴可不填）
    - local_proxy: 是否在本机启动代理服务。机器人软件 < v7.7 或无示教器时必须 True。
    """
    _ensure_sdk_ready()
    log_py(
        f"SDK Arm.connect begin controller_ip={controller_ip!r} "
        f"teach_panel_ip={teach_panel_ip!r} local_proxy={local_proxy!r}"
    )
    with redirect_sdk_stdout_to_stderr():
        try:
            arm = Arm(local_proxy=local_proxy)
        except TypeError:
            if local_proxy:
                log_py("Arm(local_proxy=...) 不受当前 SDK 版本支持，回退 Arm()")
            arm = Arm()
        if teach_panel_ip:
            ret = arm.connect(controller_ip, teach_panel_ip)
        else:
            ret = arm.connect(controller_ip)
    if ret != StatusCodeEnum.OK:
        err = getattr(ret, "errmsg", ret)
        log_py(f"SDK Arm.connect failed: {err!r}")
        gbt_raise("GBT_CONNECT_FAILED")
    log_py("SDK Arm.connect ok")
    return arm


def _check_p_service(ip: str) -> None:
    """检查 P 寄存器服务（program_pose HTTP API）端口是否可达。

    P 寄存器使用独立的 HTTP 服务（端口 {P_SERVICE_PORT}），与 R/PR 寄存器所用通道不同。
    若端口不可达，提前报出友好错误，而非让 SDK 抛出晦涩的 HTTPConnectionPool 异常。
    """
    log_py(f"_check_p_service begin ip={ip!r} port={P_SERVICE_PORT}")
    try:
        sock = socket.create_connection((ip, P_SERVICE_PORT), timeout=3.0)
        sock.close()
        log_py(f"_check_p_service ok ip={ip!r} port={P_SERVICE_PORT}")
    except OSError as exc:
        log_py(f"_check_p_service failed ip={ip!r} port={P_SERVICE_PORT} exc={exc!r}")
        gbt_raise("GBT_P_SERVICE_UNREACHABLE")


def _safe_disconnect(arm: Any) -> None:
    try:
        with redirect_sdk_stdout_to_stderr():
            arm.disconnect()
    except Exception as exc:  # pragma: no cover
        log_py(f"SDK Arm.disconnect swallow: {exc!r}")


def verify_connect(conn: Dict[str, Any]) -> Dict[str, Any]:
    log_py(
        f"verify_connect begin controller_ip={conn['controller_ip']!r} "
        f"teach_panel_ip={conn['teach_panel_ip']!r} local_proxy={conn['local_proxy']!r}"
    )
    try:
        arm = connect_arm(**conn)
    except Exception as exc:
        log_py(f"verify_connect exception: {exc!r}")
        return {"ok": False, "code": "GBT_CONNECT_FAILED", "message": str(exc)}
    _safe_disconnect(arm)
    log_py("verify_connect disconnected")
    return {"ok": True, "code": "connect_ok"}


def read_r(arm, indexes: List[int], progress: Optional[ProgressCallback] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with redirect_sdk_stdout_to_stderr():
        for pos, idx in enumerate(indexes, start=1):
            # 先上报“正在读取第 pos 条”，再实际读取；matched 为此前已匹配条数。
            if progress:
                progress(pos, len(indexes), len(rows))
            value, ret = arm.register.read_R(idx)
            if ret == StatusCodeEnum.OK:
                rows.append({"type": "R", "ID": idx, "value": round3(value)})
    return rows


def read_pr(arm, indexes: List[int], progress: Optional[ProgressCallback] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with redirect_sdk_stdout_to_stderr():
        for step, idx in enumerate(indexes, start=1):
            if progress:
                progress(step, len(indexes), len(rows))
            try:
                pose, ret = arm.register.read_PR(idx)
            except Exception as exc:
                log_py(f"read_pr skip id={idx} exc={exc!r}")
                continue
            if ret == StatusCodeEnum.OK:
                try:
                    position = pose.poseRegisterData.cartData.position
                    coord = _coord_from_pose_register(pose)
                    rows.append(
                        {
                            "TYPE": "PR",
                            "ID": idx,
                            "X": round3(position.x),
                            "Y": round3(position.y),
                            "Z": round3(position.z),
                            "A": round3(position.a),
                            "B": round3(position.b),
                            "C": round3(position.c),
                            "coord": coord,
                        }
                    )
                except Exception as exc:
                    log_py(f"read_pr skip malformed id={idx} exc={exc!r}")
    return rows


def read_p(arm, program_name: str, indexes: List[int], progress: Optional[ProgressCallback] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with redirect_sdk_stdout_to_stderr():
        for step, idx in enumerate(indexes, start=1):
            if progress:
                progress(step, len(indexes), len(rows))
            try:
                pose, ret = arm.program_pose.read(program_name, idx)
            except Exception as exc:
                log_py(f"read_p skip program={program_name!r} id={idx} exc={exc!r}")
                continue
            if ret == StatusCodeEnum.OK:
                try:
                    position = pose.poseData.cartData.baseCart.position
                    tf = _safe_int(getattr(pose.poseData.cartData, "tf", 0), 0)
                    uf = _safe_int(getattr(pose.poseData.cartData, "uf", 0), 0)
                    coord = _coord_from_pose(pose)
                    rows.append(
                        {
                            "Type": "P",
                            "ID": idx,
                            "X": round3(position.x),
                            "Y": round3(position.y),
                            "Z": round3(position.z),
                            "A": round3(position.a),
                            "B": round3(position.b),
                            "C": round3(position.c),
                            "TF": tf,
                            "UF": uf,
                            "Coord": coord,
                        }
                    )
                except Exception as exc:
                    log_py(f"read_p skip malformed program={program_name!r} id={idx} exc={exc!r}")
    return rows


def read_r_all_scan(arm, progress: Optional[ProgressCallback] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    consec_fail = 0
    idx = READ_ALL_FIRST_ID
    with redirect_sdk_stdout_to_stderr():
        while idx <= READ_ALL_MAX_ID and consec_fail < READ_ALL_CONSECUTIVE_FAIL_LIMIT:
            value, ret = arm.register.read_R(idx)
            if ret == StatusCodeEnum.OK:
                rows.append({"type": "R", "ID": idx, "value": round3(value)})
                consec_fail = 0
            else:
                consec_fail += 1
            if progress:
                progress(idx, None, len(rows))
            idx += 1
    return rows


def read_pr_all_scan(arm, progress: Optional[ProgressCallback] = None) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    consec_fail = 0
    idx = READ_ALL_FIRST_ID
    with redirect_sdk_stdout_to_stderr():
        while idx <= READ_ALL_MAX_ID and consec_fail < READ_ALL_CONSECUTIVE_FAIL_LIMIT:
            pose, ret = arm.register.read_PR(idx)
            if ret == StatusCodeEnum.OK:
                pos = pose.poseRegisterData.cartData.position
                coord = _coord_from_pose_register(pose)
                rows.append(
                    {
                        "TYPE": "PR",
                        "ID": idx,
                        "X": round3(pos.x),
                        "Y": round3(pos.y),
                        "Z": round3(pos.z),
                        "A": round3(pos.a),
                        "B": round3(pos.b),
                        "C": round3(pos.c),
                        "coord": coord,
                    }
                )
                consec_fail = 0
            else:
                consec_fail += 1
            if progress:
                progress(idx, None, len(rows))
            idx += 1
    return rows


def read_p_pose_map(
    arm,
    program_name: str,
    id_filter: Optional[Set[int]] = None,
    strict: bool = True,
) -> Dict[int, Any]:
    """Read all program poses and index them by valid pose.id."""
    with redirect_sdk_stdout_to_stderr():
        poses, ret = arm.program_pose.read_all_poses(program_name)
    if ret != StatusCodeEnum.OK:
        log_py(f"read_p_pose_map failed program={program_name!r}: {status_text(ret)}")
        if not strict:
            raise RuntimeError(status_text(ret))
        gbt_raise("GBT_INTERNAL_ERROR")
    out: Dict[int, Any] = {}
    for pose in poses:
        idx = _normalize_id(getattr(pose, "id", None))
        if idx is None:
            log_py(f"read_p_pose_map skip invalid pose id={getattr(pose, 'id', '?')!r}")
            continue
        if id_filter is not None and idx not in id_filter:
            continue
        out[idx] = pose
    return out


def read_p_all_scan(
    arm,
    program_name: str,
    progress: Optional[ProgressCallback] = None,
    id_filter: Optional[Set[int]] = None,
) -> List[Dict[str, Any]]:
    """使用 read_all_poses 一次性获取程序中所有P点（SDK 4.4.3）。"""
    pose_map = read_p_pose_map(arm, program_name, id_filter)
    rows: List[Dict[str, Any]] = []
    total = len(pose_map)
    for step, idx in enumerate(sorted(pose_map), start=1):
        pose = pose_map[idx]
        try:
            position = pose.poseData.cartData.baseCart.position
            tf = _safe_int(getattr(pose.poseData.cartData, "tf", 0), 0)
            uf = _safe_int(getattr(pose.poseData.cartData, "uf", 0), 0)
            coord = _coord_from_pose(pose)
            rows.append(
                {
                    "Type": "P",
                    "ID": idx,
                    "X": round3(position.x),
                    "Y": round3(position.y),
                    "Z": round3(position.z),
                    "A": round3(position.a),
                    "B": round3(position.b),
                    "C": round3(position.c),
                    "TF": tf,
                    "UF": uf,
                    "Coord": coord,
                }
            )
        except Exception as exc:
            log_py(f"read_p_all_scan skip pose id={getattr(pose, 'id', '?')} exc={exc!r}")
        if progress:
            progress(step, total, len(rows))
    return rows


def read_preview(conn: Dict[str, Any], req: Dict[str, Any]) -> Dict[str, Any]:
    reg_type = req["registerType"]
    selector = req.get("selector", {})
    mode = selector.get("mode", "range")
    program_name = req.get("programName")
    controller_ip = conn["controller_ip"]
    log_py(
        f"read_preview begin type={reg_type!r} controller_ip={controller_ip!r} "
        f"mode={mode!r} program={program_name!r}"
    )

    # P 寄存器使用独立 HTTP 端口，提前检查可达性，避免返回误导性的空列表。
    if reg_type == "P":
        if not program_name:
            gbt_raise("GBT_P_READ_NEED_PROGRAM")
        _check_p_service(controller_ip)

    arm = connect_arm(**conn)
    progress = lambda current, total, matched: emit_progress("read", current, total, matched)
    try:
        if mode == "all":
            if reg_type == "R":
                rows = read_r_all_scan(arm, progress)
            elif reg_type == "PR":
                rows = read_pr_all_scan(arm, progress)
            elif reg_type == "P":
                rows = read_p_all_scan(arm, program_name, progress)
            else:
                gbt_raise("GBT_UNSUPPORTED_REGISTER_TYPE")
        elif reg_type == "R":
            rows = read_r(arm, build_indexes(selector), progress)
        elif reg_type == "PR":
            rows = read_pr(arm, build_indexes(selector), progress)
        elif reg_type == "P":
            rows = read_p(arm, program_name, build_indexes(selector), progress)
        else:
            gbt_raise("GBT_UNSUPPORTED_REGISTER_TYPE")
        log_py(f"read_preview ok row_count={len(rows)}")
        return {"rows": rows}
    finally:
        _safe_disconnect(arm)
        log_py("read_preview disconnected")


def write_r(arm, row: Dict[str, Any], policy: str) -> Tuple[bool, str]:
    idx = int(row["ID"])
    value = round3(row["value"])
    with redirect_sdk_stdout_to_stderr():
        _, ret = arm.register.read_R(idx)
    exists = ret == StatusCodeEnum.OK
    if exists and policy == "skip":
        return True, "skip"
    with redirect_sdk_stdout_to_stderr():
        ret = unwrap_status(arm.register.write_R(idx, value))
    if ret != StatusCodeEnum.OK:
        return False, status_text(ret)
    return True, "write"


def build_pose_register_from_row(row: Dict[str, Any]):
    pose_register = PoseRegister()
    pose_register.id = int(row["ID"])
    pose_register.poseRegisterData.pt = PoseType.CART
    pos = pose_register.poseRegisterData.cartData.position
    pos.x = round3(row["X"])
    pos.y = round3(row["Y"])
    pos.z = round3(row["Z"])
    pos.a = round3(row["A"])
    pos.b = round3(row["B"])
    pos.c = round3(row["C"])
    coord_val = row.get("coord", row.get("coord（L/R）", "R"))
    _apply_coord_to_pose_register(pose_register, str(coord_val))
    return pose_register


def write_pr(arm, row: Dict[str, Any], policy: str) -> Tuple[bool, str]:
    idx = int(row["ID"])
    with redirect_sdk_stdout_to_stderr():
        old, ret = arm.register.read_PR(idx)
    exists = ret == StatusCodeEnum.OK
    if exists and policy == "skip":
        return True, "skip"
    if exists:
        pose_register = old
        pos = pose_register.poseRegisterData.cartData.position
        pos.x = round3(row["X"])
        pos.y = round3(row["Y"])
        pos.z = round3(row["Z"])
        pos.a = round3(row["A"])
        pos.b = round3(row["B"])
        pos.c = round3(row["C"])
        coord_val = row.get("coord", row.get("coord（L/R）", "R"))
        _apply_coord_to_pose_register(pose_register, str(coord_val))
    else:
        pose_register = build_pose_register_from_row(row)
    with redirect_sdk_stdout_to_stderr():
        ret = unwrap_status(arm.register.write_PR(pose_register))
    if ret != StatusCodeEnum.OK:
        if status_indicates_exists(ret):
            if policy == "skip":
                return True, "skip"
            return False, "ID already exists on robot"
        return False, status_text(ret)
    return True, "write"


def write_p(arm, program_name: str, row: Dict[str, Any], policy: str) -> Tuple[bool, str]:
    idx = int(row["ID"])
    try:
        pose_map = read_p_pose_map(arm, program_name, {idx}, strict=False)
    except RuntimeError as exc:
        return False, str(exc)
    pose = pose_map.get(idx)
    exists = pose is not None
    if exists and policy == "skip":
        return True, "skip"
    if exists:
        target_pose = pose
    else:
        if not hasattr(arm.program_pose, "add"):
            return False, "当前 SDK 不支持 program_pose.add，无法新增程序点。"
        if ProgramPose is None:
            return False, "未找到 ProgramPose 类型，无法新增程序点。"
        target_pose = ProgramPose()
        target_pose.id = idx
        target_pose.poseData.pt = PoseType.CART

    pos = target_pose.poseData.cartData.baseCart.position
    pos.x = round3(row["X"])
    pos.y = round3(row["Y"])
    pos.z = round3(row["Z"])
    pos.a = round3(row["A"])
    pos.b = round3(row["B"])
    pos.c = round3(row["C"])
    tf_val = row.get("TF")
    uf_val = row.get("UF")
    target_pose.poseData.cartData.tf = _safe_int(
        tf_val, _safe_int(getattr(target_pose.poseData.cartData, "tf", 0), 0)
    )
    target_pose.poseData.cartData.uf = _safe_int(
        uf_val, _safe_int(getattr(target_pose.poseData.cartData, "uf", 0), 0)
    )
    coord_val = row.get("Coord", row.get("Coord（L/R）", "R"))
    _apply_coord_to_pose(target_pose, str(coord_val))

    try:
        with redirect_sdk_stdout_to_stderr():
            if exists:
                ret = unwrap_status(arm.program_pose.write(program_name, idx, target_pose))
            else:
                ret = unwrap_status(arm.program_pose.add(program_name, idx, target_pose))
    except Exception as exc:
        exc_str = str(exc)
        if "10061" in exc_str or "HTTPConnectionPool" in exc_str or "Connection refused" in exc_str.lower():
            return False, (
                f"P寄存器服务（端口{P_SERVICE_PORT}）连接中断，写入失败。"
                f"请检查机器人控制器网络状态。"
            )
        return False, f"写入P点异常: {exc_str[:200]}"
    if ret != StatusCodeEnum.OK:
        return False, status_text(ret)
    return True, "write"


def apply_rows(conn: Dict[str, Any], req: Dict[str, Any]) -> Dict[str, Any]:
    reg_type = req["registerType"]
    policy = req.get("conflictPolicy", "skip")
    rows: List[Dict[str, Any]] = req.get("rows", [])
    program_name = req.get("programName")
    controller_ip = conn["controller_ip"]
    log_py(
        f"apply_rows begin controller_ip={controller_ip!r} type={reg_type!r} "
        f"policy={policy!r} row_count={len(rows)} program={program_name!r}"
    )

    # P 寄存器使用独立 HTTP 端口，提前检查可达性，避免写入时抛出晦涩的 SDK 异常。
    if reg_type == "P":
        if not program_name:
            gbt_raise("GBT_P_WRITE_NEED_PROGRAM")
        _check_p_service(controller_ip)

    arm = connect_arm(**conn)
    success = 0
    skipped = 0
    failed: List[str] = []
    try:
        total = len(rows)
        for pos, row in enumerate(rows, start=1):
            # 在写入当前寄存器之前先上报进度：current=正在写入的序号，matched=此前已完成的条数。
            # 这样前端会显示 “正在写入 1/10，已完成 0 条” -> “正在写入 2/10，已完成 1 条” ...
            emit_progress("write", pos, total, success + skipped)
            if reg_type == "R":
                ok, tag = write_r(arm, row, policy)
            elif reg_type == "PR":
                ok, tag = write_pr(arm, row, policy)
            elif reg_type == "P":
                ok, tag = write_p(arm, program_name, row, policy)
            else:
                gbt_raise("GBT_UNSUPPORTED_REGISTER_TYPE")

            if ok and tag == "skip":
                skipped += 1
            elif ok:
                success += 1
            else:
                failed.append(f"ID={row.get('ID')}: {tag}")
    finally:
        _safe_disconnect(arm)
        log_py("apply_rows disconnected")

    if failed:
        if (
            reg_type == "P"
            and success == 0
            and skipped == 0
            and all(("PROGRAM_NOT_FOUND" in f) or ("找不到对应的程序" in f) for f in failed)
        ):
            return {
                "ok": False,
                "code": "GBT_PROGRAM_NOT_FOUND",
                "details": [],
            }
        log_py(
            f"apply_rows end partial_fail success={success} skipped={skipped} failed={len(failed)}"
        )
        return {
            "ok": False,
            "stats": {"success": success, "skipped": skipped, "failed": len(failed)},
            "details": failed,
        }
    log_py(f"apply_rows end ok success={success} skipped={skipped}")
    return {
        "ok": True,
        "stats": {"success": success, "skipped": skipped, "failed": 0},
        "details": [],
    }


def _cell_for_export(row: Dict[str, Any], key: str, reg_type: str) -> Any:
    if key in row:
        return row[key]
    if reg_type == "P" and key == "Coord":
        return row.get("Coord（L/R）", "")
    if reg_type == "PR" and key == "coord":
        return row.get("coord（L/R）", "")
    return row.get(key, "")


def export_excel(register_type: str, rows: List[Dict[str, Any]], output_path: str) -> Dict[str, Any]:
    if Workbook is None:
        if _OPENPYXL_ERR is not None:
            log_py(f"openpyxl missing: {_OPENPYXL_ERR!r}")
        gbt_raise("GBT_OPENPYXL_MISSING")
    log_py(f"export_excel begin type={register_type!r} rows={len(rows)} path_len={len(output_path)}")
    headers = make_headers(register_type)
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # 超过 5000 行启用流式写入（write_only），避免整表驻留内存。
    use_write_only = len(rows) > 5000
    wb = Workbook(write_only=use_write_only)
    if use_write_only:
        ws = wb.create_sheet(title=register_type)
        ws.append(headers)
        for row in rows:
            record = [_cell_for_export(row, key, register_type) for key in headers]
            record = [round3(v) if isinstance(v, (int, float)) else v for v in record]
            ws.append(record)
    else:
        ws = wb.active
        ws.title = register_type
        ws.append(headers)
        for row in rows:
            record = [_cell_for_export(row, key, register_type) for key in headers]
            record = [round3(v) if isinstance(v, (int, float)) else v for v in record]
            ws.append(record)

    wb.save(output_file)
    log_py(f"export_excel ok path={output_file} write_only={use_write_only}")
    return {"ok": True, "code": "export_saved"}


def fetch_robot_meta(conn: Dict[str, Any]) -> Dict[str, Any]:
    log_py(
        f"fetch_robot_meta begin controller_ip={conn['controller_ip']!r} "
        f"teach_panel_ip={conn['teach_panel_ip']!r} local_proxy={conn['local_proxy']!r}"
    )
    arm = connect_arm(**conn)
    try:
        with redirect_sdk_stdout_to_stderr():
            model_info, ret_m = arm.get_arm_model_info()
            ver_info, ret_v = arm.get_controller_version()
        model_str = (
            str(model_info) if ret_m == StatusCodeEnum.OK and model_info is not None else ""
        )
        ver_str = str(ver_info) if ret_v == StatusCodeEnum.OK and ver_info is not None else ""
        log_py(f"fetch_robot_meta ok model_len={len(model_str)} ver_len={len(ver_str)}")
        return {"model": model_str, "controllerVersion": ver_str}
    finally:
        _safe_disconnect(arm)
        log_py("fetch_robot_meta disconnected")


def export_template(register_type: str, output_path: str) -> Dict[str, Any]:
    if Workbook is None:
        if _OPENPYXL_ERR is not None:
            log_py(f"openpyxl missing: {_OPENPYXL_ERR!r}")
        gbt_raise("GBT_OPENPYXL_MISSING")
    log_py(f"export_template begin type={register_type!r} path_len={len(output_path)}")
    wb = Workbook()
    ws = wb.active
    ws.title = register_type
    ws.append(make_headers(register_type))
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    log_py(f"export_template ok path={output_file}")
    return {"ok": True, "code": "template_saved"}


# ---------- 入口 ------------------------------------------------------------


def _dispatch(payload: Dict[str, Any]) -> Dict[str, Any]:
    action = payload.get("action")
    if action == "read_preview":
        return read_preview(_extract_connect_params(payload), payload["request"])
    if action == "apply_rows":
        return apply_rows(_extract_connect_params(payload), payload["request"])
    if action == "export_excel":
        return export_excel(
            payload["registerType"], payload.get("rows", []), payload["outputPath"]
        )
    if action == "export_template":
        return export_template(payload["registerType"], payload["outputPath"])
    if action == "fetch_robot_meta":
        return fetch_robot_meta(_extract_connect_params(payload))
    if action == "verify_connect":
        return verify_connect(_extract_connect_params(payload))
    gbt_raise("GBT_INTERNAL_ERROR")


def main() -> int:
    try:
        payload = read_payload()
    except Exception as exc:
        log_py(f"read_payload failed: {exc!r}")
        emit_frame({"ok": False, "code": "GBT_INTERNAL_ERROR", "message": "GBT_INTERNAL_ERROR"})
        return 0

    action = payload.get("action", "?")
    log_py(f"main action={action!r} keys={list(payload.keys())}")
    try:
        result = _dispatch(payload)
    except Exception as exc:
        code = sidecar_error_code(exc)
        if code == "GBT_INTERNAL_ERROR":
            log_py(f"main action={action!r} exception: {exc!r}\n{traceback.format_exc()}")
        emit_frame({"ok": False, "code": code, "message": code})
        return 0

    if not isinstance(result, dict):
        emit_frame(
            {
                "ok": False,
                "code": "GBT_INTERNAL_ERROR",
                "message": "GBT_INTERNAL_ERROR",
            }
        )
        return 0

    log_py(f"main action={action!r} done ok_field={result.get('ok')!r}")
    emit_frame(result)
    return 0


if __name__ == "__main__":
    sys.exit(main())
